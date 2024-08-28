from django.shortcuts import redirect, render
from django.http import HttpResponse,HttpResponseRedirect
from employee_information.models import Department, Position, Employees,Course,Student_bio,Announcement,MaintenanceTask,Equipment,Attendance
from django.contrib.auth import authenticate, login, logout
from django.contrib.auth.decorators import login_required
from django.contrib.auth.hashers import check_password  
from django.shortcuts import get_object_or_404
from django.contrib import messages
from django.shortcuts import redirect
from .forms import StudentSignupForm,MaintenanceTaskForm,EmployeeLoginForm,StudentLoginForm
from django.views.decorators.csrf import  csrf_exempt
from .models import Student,Employees
from django.http import JsonResponse
from django.core.mail import send_mail
from django.conf import settings
from django.urls import reverse
import pandas as pd
import requests
import openpyxl
import json
from .forms import MaintenanceTaskForm
from openpyxl.utils import get_column_letter
import logging

logger = logging.getLogger(__name__)

####################################################

#employee login

def login_emp(request):
    if request.method == 'POST':
        form = EmployeeLoginForm(request.POST)
        if form.is_valid():
            email = form.cleaned_data['email']
            code = form.cleaned_data['code']
            try:
                employee = Employees.objects.get(email=email, code=code)
                # Assuming you have a way to authenticate employees
                request.session['employee_id'] = employee.id  # Store employee ID in session
                return redirect('home-emp')  # Redirect to the home page after login
            except Employees.DoesNotExist:
                messages.error(request, 'Invalid email or code')
    else:
        form = EmployeeLoginForm()
    return render(request, 'teacher_information/login_emp.html', {'form': form})
@login_required
def logout_emp(request):
    # Clear all session data
    for key in list(request.session.keys()):
        del request.session[key]
    
    # Log out the user
    logout(request)
    
    # Redirect to login page
    return redirect('login_emp')

@login_required
def home_emp(request):
    # Ensure the user is authenticated; this is a simple example, and you may need more robust checks
    if not request.session.get('employee_id'):
        return redirect('login_emp')
    total_courses = Course.objects.count()  
    return render(request, 'teacher_information/home_emp.html', {'total_courses': total_courses})
#####################################################

#maintanance task


@login_required
def maintenance_tasks(request):
    task_list = MaintenanceTask.objects.all()
    context = {
        'page_title': 'Maintenance Tasks',
        'maintenance_tasks': task_list,
    }
    return render(request, 'employee_information/maintainance.html', context)

@login_required
def manage_maintenance_tasks(request):
    task = {}
    equipment_list = Equipment.objects.all()  # To populate the equipment dropdown in the form
    if request.method == 'GET':
        data = request.GET
        id = ''
        if 'id' in data:
            id = data['id']
        if id.isnumeric() and int(id) > 0:
            task = MaintenanceTask.objects.filter(id=id).first()
    
    context = {
        'task': task,
        'equipment_list': equipment_list
    }
    return render(request, 'employee_information/manage_maintainance.html', context)

@login_required
def save_maintenance_task(request):
    data = request.POST
    resp = {'status': 'failed'}
    try:
        equipment = Equipment.objects.get(id=data['equipment'])
        if data['id'].isnumeric() and int(data['id']) > 0:
            save_task = MaintenanceTask.objects.filter(id=data['id']).update(
                equipment=equipment,
                task_name=data['task_name'],
                scheduled_date=data['scheduled_date'],
                completed=data['completed'] == 'True'
            )
        else:
            save_task = MaintenanceTask(
                equipment=equipment,
                task_name=data['task_name'],
                scheduled_date=data['scheduled_date'],
                completed=data['completed'] == 'True'
            )
            save_task.save()
        resp['status'] = 'success'
    except Exception as e:
        print(e)
        resp['status'] = 'failed'
    return HttpResponse(json.dumps(resp), content_type="application/json")

@login_required
def delete_maintenance_task(request):
    data = request.POST
    resp = {'status': ''}
    try:
        MaintenanceTask.objects.filter(id=data['id']).delete()
        resp['status'] = 'success'
    except Exception as e:
        print(e)
        resp['status'] = 'failed'
    return HttpResponse(json.dumps(resp), content_type="application/json")

########################################################################

employees = [

    {
        'code':1,
        'name':"John D Smith",
        'contact':'09123456789',
        'address':'Sample Address only'
    },{
        'code':2,
        'name':"Claire C Blake",
        'contact':'09456123789',
        'address':'Sample Address2 only'
    }

]
# Login
def login_user(request):
    logout(request)
    resp = {"status":'failed','msg':''}
    username = ''
    password = ''
    if request.POST:
        username = request.POST['username']
        password = request.POST['password']

        user = authenticate(username=username, password=password)
        if user is not None:
            if user.is_active:
                login(request, user)
                resp['status']='success'
            else:
                resp['msg'] = "Incorrect username or password"
        else:
            resp['msg'] = "Incorrect username or password"
    return HttpResponse(json.dumps(resp),content_type='application/json')

#Logout
def logoutuser(request):
    logout(request)
    return redirect('/')

##################################################################

# Create your views here.
@login_required
def home(request):
    context = {
        'page_title':'Home',
        'employees':employees,
        'total_department':len(Department.objects.all()),
        'total_position':len(Position.objects.all()),
        'total_employee':len(Employees.objects.all()),
        'total_student':len(Student_bio.objects.all()),
        'total_course':len(Course.objects.all())
    }
    return render(request, 'employee_information/home.html',context)


def about(request):
    context = {
        'page_title':'About',
    }
    return render(request, 'employee_information/about.html',context)
######################################################################################


# Departments
@login_required
def export_departments_to_excel(request):
    # Create a workbook and a worksheet
    workbook = openpyxl.Workbook()
    worksheet = workbook.active
    worksheet.title = 'Departments'

    # Define the header row
    headers = ['ID', 'Department Name', 'Description', 'Status']
    worksheet.append(headers)

    # Add department data
    for department in Department.objects.all():
        worksheet.append([
            department.id,
            department.name,
            department.description,
            'Active' if department.status == 1 else 'Inactive'
        ])

    # Set the response
    response = HttpResponse(content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
    response['Content-Disposition'] = 'attachment; filename=departments.xlsx'
    
    # Save the workbook to the response
    workbook.save(response)
    return response
@login_required
def departments(request):
    department_list = Department.objects.all()
    context = {
        'page_title':'Departments',
        'departments':department_list,
    }
    return render(request, 'employee_information/departments.html',context)
@login_required
def manage_departments(request):
    department = {}
    if request.method == 'GET':
        data =  request.GET
        id = ''
        if 'id' in data:
            id= data['id']
        if id.isnumeric() and int(id) > 0:
            department = Department.objects.filter(id=id).first()
    
    context = {
        'department' : department
    }
    return render(request, 'employee_information/manage_department.html',context)

@login_required
def save_department(request):
    data =  request.POST
    resp = {'status':'failed'}
    try:
        if (data['id']).isnumeric() and int(data['id']) > 0 :
            save_department = Department.objects.filter(id = data['id']).update(name=data['name'], description = data['description'],status = data['status'])
        else:
            save_department = Department(name=data['name'], description = data['description'],status = data['status'])
            save_department.save()
        resp['status'] = 'success'
    except:
        resp['status'] = 'failed'
    return HttpResponse(json.dumps(resp), content_type="application/json")

@login_required
def delete_department(request):
    data =  request.POST
    resp = {'status':''}
    try:
        Department.objects.filter(id = data['id']).delete()
        resp['status'] = 'success'
    except:
        resp['status'] = 'failed'
    return HttpResponse(json.dumps(resp), content_type="application/json")


##############################################################################################


# Positions
@login_required
def positions(request):
    position_list = Position.objects.all()
    context = {
        'page_title':'Positions',
        'positions':position_list,
    }
    return render(request, 'employee_information/positions.html',context)
@login_required
def manage_positions(request):
    position = {}
    if request.method == 'GET':
        data =  request.GET
        id = ''
        if 'id' in data:
            id= data['id']
        if id.isnumeric() and int(id) > 0:
            position = Position.objects.filter(id=id).first()
    
    context = {
        'position' : position
    }
    return render(request, 'employee_information/manage_position.html',context)

@login_required
def save_position(request):
    data =  request.POST
    resp = {'status':'failed'}
    try:
        if (data['id']).isnumeric() and int(data['id']) > 0 :
            save_position = Position.objects.filter(id = data['id']).update(name=data['name'], description = data['description'],status = data['status'])
        else:
            save_position = Position(name=data['name'], description = data['description'],status = data['status'])
            save_position.save()
        resp['status'] = 'success'
    except:
        resp['status'] = 'failed'
    return HttpResponse(json.dumps(resp), content_type="application/json")

@login_required
def delete_position(request):
    data =  request.POST
    resp = {'status':''}
    try:
        Position.objects.filter(id = data['id']).delete()
        resp['status'] = 'success'
    except:
        resp['status'] = 'failed'
    return HttpResponse(json.dumps(resp), content_type="application/json")
@login_required
def export_positions_to_excel(request):
    # Create a workbook and a worksheet
    workbook = openpyxl.Workbook()
    worksheet = workbook.active
    worksheet.title = 'Positions'

    # Define the header row
    headers = ['ID', 'Position Name', 'Description', 'Status']
    worksheet.append(headers)

    # Add position data
    for position in Position.objects.all():
        worksheet.append([
            position.id,
            position.name,
            position.description,
            'Active' if position.status == 1 else 'Inactive'
        ])

    # Set the response
    response = HttpResponse(content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
    response['Content-Disposition'] = 'attachment; filename=positions.xlsx'
    
    # Save the workbook to the response
    workbook.save(response)
    return response


##################################################################################################


#courses
@login_required
def export_courses_to_excel(request):
    # Create a workbook and a worksheet
    workbook = openpyxl.Workbook()
    worksheet = workbook.active
    worksheet.title = 'Courses'

    # Define the header row
    headers = ['ID', 'Name', 'Description', 'Status']
    worksheet.append(headers)

    # Add course data
    for course in Course.objects.all():
        worksheet.append([
            course.id,
            course.name,
            course.description,
            'Active' if course.status == 1 else 'Inactive'
        ])

    # Set the response
    response = HttpResponse(content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
    response['Content-Disposition'] = 'attachment; filename=courses.xlsx'
    
    # Save the workbook to the response
    workbook.save(response)
    return response
def courses(request):
    course_list = Course.objects.all()
    context = {
        'page_title': 'Courses',
        'courses': course_list,
    }
    return render(request, 'employee_information/courses.html', context)

@login_required
def manage_courses(request):
    course = {}
    if request.method == 'GET':
        data = request.GET
        id = ''
        if 'id' in data:
            id = data['id']
        if id.isnumeric() and int(id) > 0:
            course = Course.objects.filter(id=id).first()
    
    context = {
        'course': course
    }
    return render(request, 'employee_information/manage_course.html', context)

@login_required
def save_course(request):
    data = request.POST
    resp = {'status': 'failed'}
    try:
        if (data['id']).isnumeric() and int(data['id']) > 0:
            Course.objects.filter(id=data['id']).update(name=data['name'], description=data['description'], status=data['status'])
        else:
            new_course = Course(name=data['name'], description=data['description'], status=data['status'])
            new_course.save()
        resp['status'] = 'success'
    except Exception as e:
        resp['status'] = 'failed'
        resp['error'] = str(e)
    return HttpResponse(json.dumps(resp), content_type="application/json")

@login_required
def delete_course(request):
    data = request.POST
    resp = {'status': ''}
    try:
        Course.objects.filter(id=data['id']).delete()
        resp['status'] = 'success'
    except Exception as e:
        resp['status'] = 'failed'
        resp['error'] = str(e)
    return HttpResponse(json.dumps(resp), content_type="application/json")


##################################################################################

@login_required
# Employees
def export_employees_to_excel(request):
    # Create a workbook and a worksheet
    workbook = openpyxl.Workbook()
    worksheet = workbook.active
    worksheet.title = 'Employees'

    # Define the header row
    headers = [
        'ID', 'Code', 'First Name', 'Middle Name', 'Last Name', 'Gender', 
        'Date of Birth', 'Contact', 'Address', 'Email', 'Department', 
        'Position', 'Date Hired', 'Salary', 'Status', 'Date Added', 'Date Updated'
    ]
    worksheet.append(headers)

    # Add employee data
    for employee in Employees.objects.all():
        worksheet.append([
            employee.id,
            employee.code,
            employee.firstname,
            employee.middlename,
            employee.lastname,
            employee.gender,
            employee.dob,
            employee.contact,
            employee.address,
            employee.email,
            employee.department_id.name if employee.department_id else '',
            employee.position_id.name if employee.position_id else '',
            employee.date_hired,
            employee.salary,
            'Active' if employee.status == 1 else 'Inactive',
            employee.date_added,
            employee.date_updated
        ])

    # Set the response
    response = HttpResponse(content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
    response['Content-Disposition'] = 'attachment; filename=employees.xlsx'
    
    # Save the workbook to the response
    workbook.save(response)
    return response

def employees(request):
    employee_list = Employees.objects.all()
    context = {
        'page_title':'Employees',
        'employees':employee_list,
    }
    return render(request, 'employee_information/employees.html',context)
@login_required
def manage_employees(request):
    employee = {}
    departments = Department.objects.filter(status = 1).all() 
    positions = Position.objects.filter(status = 1).all() 
    if request.method == 'GET':
        data =  request.GET
        id = ''
        if 'id' in data:
            id= data['id']
        if id.isnumeric() and int(id) > 0:
            employee = Employees.objects.filter(id=id).first()
    context = {
        'employee' : employee,
        'departments' : departments,
        'positions' : positions
    }
    return render(request, 'employee_information/manage_employee.html',context)

@login_required
def save_employee(request):
    data =  request.POST
    resp = {'status':'failed'}
    if (data['id']).isnumeric() and int(data['id']) > 0:
        check  = Employees.objects.exclude(id = data['id']).filter(code = data['code'])
    else:
        check  = Employees.objects.filter(code = data['code'])

    if len(check) > 0:
        resp['status'] = 'failed'
        resp['msg'] = 'Code Already Exists'
    else:
        try:
            dept = Department.objects.filter(id=data['department_id']).first()
            pos = Position.objects.filter(id=data['position_id']).first()
            if (data['id']).isnumeric() and int(data['id']) > 0 :
                save_employee = Employees.objects.filter(id = data['id']).update(code=data['code'], firstname = data['firstname'],middlename = data['middlename'],lastname = data['lastname'],dob = data['dob'],gender = data['gender'],contact = data['contact'],email = data['email'],address = data['address'],department_id = dept,position_id = pos,date_hired = data['date_hired'],salary = data['salary'],status = data['status'])
            else:
                save_employee = Employees(code=data['code'], firstname = data['firstname'],middlename = data['middlename'],lastname = data['lastname'],dob = data['dob'],gender = data['gender'],contact = data['contact'],email = data['email'],address = data['address'],department_id = dept,position_id = pos,date_hired = data['date_hired'],salary = data['salary'],status = data['status'])
                save_employee.save()
            resp['status'] = 'success'
        except Exception:
            resp['status'] = 'failed'
            print(Exception)
            print(json.dumps({"code":data['code'], "firstname" : data['firstname'],"middlename" : data['middlename'],"lastname" : data['lastname'],"dob" : data['dob'],"gender" : data['gender'],"contact" : data['contact'],"email" : data['email'],"address" : data['address'],"department_id" : data['department_id'],"position_id" : data['position_id'],"date_hired" : data['date_hired'],"salary" : data['salary'],"status" : data['status']}))
    return HttpResponse(json.dumps(resp), content_type="application/json")

@login_required
def delete_employee(request):
    data =  request.POST
    resp = {'status':''}
    try:
        Employees.objects.filter(id = data['id']).delete()
        resp['status'] = 'success'
    except:
        resp['status'] = 'failed'
    return HttpResponse(json.dumps(resp), content_type="application/json")

@login_required
def view_employee(request):
    employee = {}
    departments = Department.objects.filter(status = 1).all() 
    positions = Position.objects.filter(status = 1).all() 
    if request.method == 'GET':
        data =  request.GET
        id = ''
        if 'id' in data:
            id= data['id']
        if id.isnumeric() and int(id) > 0:
            employee = Employees.objects.filter(id=id).first()
    context = {
        'employee' : employee,
        'departments' : departments,
        'positions' : positions
    }
    return render(request, 'employee_information/view_employee.html',context)
@login_required
def charts_page(request):
    departments = Department.objects.all()
    department_names = [department.name for department in departments]
    employee_counts = [Employees.objects.filter(department_id=department.id).count() for department in departments]

    context = {
        'page_title': 'Charts',
        'department_names': json.dumps(department_names),  # JSON encode the data
        'employee_counts': json.dumps(employee_counts),  # JSON encode the data
    }
    return render(request, 'employee_information/charts.html', context)

#########################################################################################

#student bio


def students(request):
    student_list = Student_bio.objects.all()
    context = {
        'page_title': 'Students',
        'students': student_list,
    }
    return render(request, 'employee_information/students.html', context)

@login_required
def manage_students(request):
    student = {}
    departments = Department.objects.filter(status=1).all()
    courses = Course.objects.filter(status=1).all()
    if request.method == 'GET':
        data = request.GET
        id = ''
        if 'id' in data:
            id = data['id']
        if id.isnumeric() and int(id) > 0:
            student = Student_bio.objects.filter(id=id).first()
    context = {
        'student': student,
        'departments': departments,
        'courses': courses
    }
    return render(request, 'employee_information/manage_student.html', context)

@login_required
def save_student(request):
    data = request.POST
    resp = {'status': 'failed'}
    if (data['id']).isnumeric() and int(data['id']) > 0:
        check = Student_bio.objects.exclude(id=data['id']).filter(student_bio_id=data['student_bio_id'])
    else:
        check = Student_bio.objects.filter(student_bio_id=data['student_bio_id'])

    if len(check) > 0:
        resp['status'] = 'failed'
        resp['msg'] = 'Student ID Already Exists'
    else:
        try:
            dept = Department.objects.filter(id=data['department']).first()
            course = Course.objects.filter(id=data['course']).first()
            if (data['id']).isnumeric() and int(data['id']) > 0:
                Student_bio.objects.filter(id=data['id']).update(
                    student_bio_id=data['student_bio_id'],
                    first_name=data['first_name'],
                    middle_name=data['middle_name'],
                    last_name=data['last_name'],
                    dob=data['dob'],
                    gender=data['gender'],
                    contact_number=data['contact_number'],
                    email=data['email'],
                    address=data['address'],
                    department=dept,
                    course=course,
                    date_enrolled=data['date_enrolled'],
                    status=data['status']
                )
            else:
                new_student = Student_bio(
                    student_bio_id=data['student_bio_id'],
                    first_name=data['first_name'],
                    middle_name=data['middle_name'],
                    last_name=data['last_name'],
                    dob=data['dob'],
                    gender=data['gender'],
                    contact_number=data['contact_number'],
                    email=data['email'],
                    address=data['address'],
                    department=dept,
                    course=course,
                    date_enrolled=data['date_enrolled'],
                    status=data['status']
                )
                new_student.save()
            resp['status'] = 'success'
        except Exception as e:
            resp['status'] = 'failed'
            print(e)
            print(json.dumps({
                "student_bio_id": data['student_bio_id'],
                "first_name": data['first_name'],
                "middle_name": data['middle_name'],
                "last_name": data['last_name'],
                "dob": data['dob'],
                "gender": data['gender'],
                "contact_number": data['contact_number'],
                "email": data['email'],
                "address": data['address'],
                "department": data['department'],
                "course": data['course'],
                "date_enrolled": data['date_enrolled'],
                "status": data['status']
            }))
    return HttpResponse(json.dumps(resp), content_type="application/json")

@login_required
def delete_student(request):
    data = request.POST
    resp = {'status': ''}
    try:
        Student_bio.objects.filter(id=data['id']).delete()
        resp['status'] = 'success'
    except Exception as e:
        resp['status'] = 'failed'
        print(e)
    return HttpResponse(json.dumps(resp), content_type="application/json")

@login_required
def view_student(request):
    student = {}
    departments = Department.objects.filter(status=1).all()
    courses = Course.objects.filter(status=1).all()
    if request.method == 'GET':
        data = request.GET
        id = ''
        if 'id' in data:
            id = data['id']
        if id.isnumeric() and int(id) > 0:
            student = Student_bio.objects.filter(id=id).first()
    context = {
        'student': student,
        'departments': departments,
        'courses': courses
    }
    return render(request, 'employee_information/view_student.html', context)

###################################################################################################


# Announcement
@login_required 
def announcements(request):
    announcement_list = Announcement.objects.all()
    context = {
        'page_title': 'Announcements',
        'announcements': announcement_list,
    }
    return render(request, 'teacher_information/announcement.html', context)

@login_required
def manage_announcement(request):
    announcement = {}
    if request.method == 'GET':
        data = request.GET
        id = ''
        if 'id' in data:
            id = data['id']
        if id.isnumeric() and int(id) > 0:
            announcement = Announcement.objects.filter(id=id).first()
    
    context = {
        'announcement': announcement
    }
    return render(request, 'teacher_information/manage_position.html', context)

@login_required
def save_announcement(request):
    data = request.POST
    resp = {'status': 'failed'}
    try:
        if data.get('id') and data['id'].isnumeric() and int(data['id']) > 0:
            Announcement.objects.filter(id=data['id']).update(
                title=data['title'],
                description=data['description']
            )
        else:
            Announcement.objects.create(
                title=data['title'],
                description=data['description']
            )
        resp['status'] = 'success'
    except Exception as e:
        print(e)  # Print exception for debugging purposes
        resp['status'] = 'failed'
    return JsonResponse(resp)

@login_required
def delete_announcement(request):
    data = request.POST
    resp = {'status': 'failed'}
    try:
        announcement_id = data.get('id')
        if announcement_id and announcement_id.isnumeric():
            Announcement.objects.filter(id=announcement_id).delete()
            resp['status'] = 'success'
        else:
            resp['status'] = 'failed'
    except Exception as e:
        print(e)  # Print exception for debugging purposes
        resp['status'] = 'failed'
    return JsonResponse(resp)
@login_required
def save_announcement(request):
    data = request.POST
    resp = {'status': 'failed'}
    try:
        # Create or update the announcement
        if data.get('id') and data['id'].isnumeric() and int(data['id']) > 0:
            Announcement.objects.filter(id=data['id']).update(
                title=data['title'],
                description=data['description']
            )
            announcement = Announcement.objects.get(id=data['id'])
        else:
            announcement = Announcement.objects.create(
                title=data['title'],
                description=data['description']
            )

        # Get all student email addresses
        students = list(Student_bio.objects.values_list('email', flat=True))
        print("Student emails:", students)  # Print email addresses for debugging

        # Fixed subject
        subject = "New Announcement"
        message = f"Title: {announcement.title}\n\n{announcement.description}"
        from_email = settings.DEFAULT_FROM_EMAIL
        
        # Send email to all students
        send_mail(
            subject,
            message,
            from_email,
            students,
            fail_silently=False,
        )
        
        resp['status'] = 'success'
    except Exception as e:
        print("Error:", e)  # Print exception for debugging purposes
        resp['status'] = 'failed'
    return JsonResponse(resp)


##################################################################################################################

#attendace


def attendance_list(request):
    attendance_records = Attendance.objects.all()
    context = {
        'page_title': 'Attendance Records',
        'attendance_list': attendance_records,
    }
    return render(request, 'teacher_information/attendance.html', context)

@login_required
def manage_attendance(request):
    attendance = {}
    if request.method == 'GET':
        data = request.GET
        id = ''
        if 'id' in data:
            id = data['id']
        if id.isnumeric() and int(id) > 0:
            attendance = Attendance.objects.filter(id=id).first()
    
    students = Student_bio.objects.all()
    context = {
        'attendance': attendance,
        'students': students
    }
    return render(request, 'teacher_information/manage_attendace.html', context)

@login_required
def save_attendance(request):
    data = request.POST
    resp = {'status': 'failed'}
    try:
        if data.get('id') and data['id'].isnumeric() and int(data['id']) > 0:
            Attendance.objects.filter(id=data['id']).update(
                student_id=data['student'],
                date=data['date'],
                status=data['status']
            )
        else:
            Attendance.objects.create(
                student_id=data['student'],
                date=data['date'],
                status=data['status']
            )
        resp['status'] = 'success'
    except Exception as e:
        print(e)  # Print exception for debugging purposes
        resp['status'] = 'failed'
    return JsonResponse(resp)

@login_required
def delete_attendance(request):
    data = request.POST
    resp = {'status': 'failed'}
    try:
        attendance_id = data.get('id')
        if attendance_id and attendance_id.isnumeric():
            Attendance.objects.filter(id=attendance_id).delete()
            resp['status'] = 'success'
        else:
            resp['status'] = 'failed'
    except Exception as e:
        print(e)  # Print exception for debugging purposes
        resp['status'] = 'failed'
    return JsonResponse(resp)


from .utils import fetch_educational_news
def news_list(request):
    news_articles = fetch_educational_news()
    return render(request, 'teacher_information/news_list.html', {'news_articles': news_articles})

####################################################################################


## STUDENT MODULES
def student_signup(request):
    if request.method == 'POST':
        form = StudentSignupForm(request.POST)
        if form.is_valid():
            form.save()
            messages.success(request, 'Signup successful.')
            return redirect('home-page')
        else:
            messages.error(request, 'Please correct the errors below.')
    else:
        form = StudentSignupForm()
    
    return render(request, 'student_information/student_login.html', {'form': form})


@csrf_exempt
def student_login(request):
    if request.method == 'POST':
        form = StudentLoginForm(request.POST)
        if form.is_valid():
            email = form.cleaned_data['email']
            password = form.cleaned_data['password']

            try:
                student = Student.objects.get(email=email)
                if check_password(password, student.password):
                    return JsonResponse({'status': 'success', 'redirect_url': reverse('home-page')})
                else:
                    return JsonResponse({'status': 'failed', 'message': 'Invalid password'})
            except Student.DoesNotExist:
                return JsonResponse({'status': 'failed', 'message': 'Email not found'})

    else:
        form = StudentLoginForm()

    return render(request, 'student_information/student_login.html', {'form': form})
@login_required
def home_page(request):
    total_courses = Course.objects.count()
    total_depa = Department.objects.count()

    # # Fetch the quote of the day from the API
    # response = requests.get('https://favqs.com/api/qotd')
    # if response.status_code == 200:
    #     quote_data = response.json()
    #     quote = quote_data.get('quote', {}).get('body', '')
    #     author = quote_data.get('quote', {}).get('author', '')
    # else:
    #     quote = "Unable to fetch quote at this moment."
    #     author = ""

    return render(request, 'student_information/home_page_student.html', {
        'total_courses': total_courses,
        'total_depa': total_depa,
        # 'quote': quote,
        # 'author': author,
    })

@login_required
def course_list(request):
    courses = Course.objects.all()
    return render(request, 'student_information/courses-list.html', {'courses': courses})
@login_required
def announcement_list(request):
    announcements = Announcement.objects.all()
    return render(request, 'student_information/announcement-list.html', {'announcements': announcements})
@login_required
def department_list(request):
    departments = Department.objects.all()  # Fetch all department instances
    return render(request, 'student_information/department-list.html', {'departments': departments})
@login_required
def eudcation_news_list(request):
    edu_news_articles = fetch_educational_news()
    return render(request, 'student_information/news_list.html', {'edu_news_articles': edu_news_articles})
@login_required
def student_logout(request):
    logout(request)
    return redirect('student_informationstudent_login.html')





# from django.shortcuts import render

# def landing_page(request):
#     return render(request, 'landingpage.html')